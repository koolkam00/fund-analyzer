from __future__ import annotations

import io
from typing import Dict, Tuple, Optional

import pandas as pd
import streamlit as st


def ensure_workbook_loaded() -> Tuple[Dict[str, pd.DataFrame], Optional[str], Optional[str], Optional[str]]:
    """Ensure a single Excel workbook is loaded into session state for all non-PME pages.

    Returns (sheets, ops_sheet_name, funds_sheet_name, bench_sheet_name)
    """
    with st.sidebar:
        st.markdown("**Workbook**")
        header_row_index = int(st.session_state.get("header_row_index", 2))
        header_row_index = st.number_input(
            "Header row (1-based) for uploaded sheets",
            min_value=1,
            max_value=100,
            value=header_row_index,
            step=1,
            key="wb_header_row_index",
        )
        upload = st.file_uploader("Upload Excel workbook (.xlsx)", type=["xlsx"], key="wb_file_uploader")  # type: ignore
        # Provide master workbook template download under file format controls
        # Master workbook template download (headers on row 2, instructions on row 1)
        try:
            templ_bytes = build_master_workbook_template()
        except Exception:
            templ_bytes = None
        if templ_bytes:
            st.download_button(
                label="Download Master Workbook Template",
                data=templ_bytes,
                file_name="PE_Fund_Analyzer_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # If already loaded and no new upload, reuse existing
    if "workbook" in st.session_state and upload is None:
        wb = st.session_state["workbook"]
        # Render manager banner if available
        _maybe_render_manager_banner(wb.get("manager_profile"))
        return wb["sheets"], wb.get("ops_sheet_name"), wb.get("funds_sheet_name"), wb.get("bench_sheet_name")

    if upload is None:
        return {}, None, None, None

    # Read all sheets with selected header row
    header_zero = max(0, int(header_row_index) - 1)
    content = upload.getvalue()
    bio = io.BytesIO(content)
    xls = pd.ExcelFile(bio, engine="openpyxl")
    sheets: Dict[str, pd.DataFrame] = {sheet: xls.parse(sheet, header=header_zero) for sheet in xls.sheet_names}

    sheet_names = list(sheets.keys())
    ops_sheet_name = sheet_names[0] if sheet_names else None
    funds_sheet_name = sheet_names[1] if len(sheet_names) > 1 else None
    bench_sheet_name = sheet_names[2] if len(sheet_names) > 2 else None

    # Detect manager profile sheet (prefer by name, else 5th sheet)
    manager_profile = None
    mgr_sheet_name = None
    for nm in sheet_names:
        low = str(nm).strip().lower()
        if low in {"fund manager profile", "manager profile", "fund manager"}:
            mgr_sheet_name = nm
            break
    if mgr_sheet_name is None and len(sheet_names) > 4:
        mgr_sheet_name = sheet_names[4]
    if mgr_sheet_name and mgr_sheet_name in sheets:
        try:
            df_mgr = sheets[mgr_sheet_name].copy()
            if not df_mgr.empty:
                # Normalize columns and take first non-empty row as profile
                def _norm(s: str) -> str:
                    return str(s).strip().lower().replace(" ", "_")
                df_mgr.columns = [_norm(c) for c in df_mgr.columns]
                first = df_mgr.iloc[0].to_dict()
                manager_profile = {k: first.get(k) for k in df_mgr.columns}
                st.session_state["manager_profile"] = manager_profile
                # Derive firm display name (prefer firm_name)
                firm_val = None
                for key in ("firm_name", "name", "manager_name"):
                    if key in manager_profile and pd.notna(manager_profile[key]):
                        firm_val = str(manager_profile[key])
                        break
                if firm_val:
                    st.session_state["firm_name"] = firm_val
        except Exception:
            manager_profile = None

    st.session_state["workbook"] = {
        "sheets": sheets,
        "ops_sheet_name": ops_sheet_name,
        "funds_sheet_name": funds_sheet_name,
        "bench_sheet_name": bench_sheet_name,
        "header_row_index": header_row_index,
        "manager_profile": manager_profile,
    }
    # Back-compat for existing pages
    st.session_state["sheets"] = sheets
    st.session_state["selected_sheet"] = ops_sheet_name
    st.session_state["header_row_index"] = header_row_index

    # Render manager banner on main area
    _maybe_render_manager_banner(manager_profile)

    return sheets, ops_sheet_name, funds_sheet_name, bench_sheet_name


# ===== Master workbook template (Portfolio Metrics, Funds Net Returns, Benchmarks) =====
@st.cache_data(show_spinner=False)
def build_master_workbook_template() -> bytes:
    """Create an Excel workbook with:
    - Sheet "Portfolio Metrics": instructions in row 1, headers in row 2 (55 columns)
    - Sheet "Funds Net Returns": instructions in row 1, headers in row 2
    - Sheet "Benchmarks": instructions in row 1, headers in row 2

    The template aligns with the expected fixed-order columns referenced across the app.
    """
    import io
    from openpyxl.utils import get_column_letter
    
    # Portfolio Metrics headers (ordered 1..55)
    portfolio_headers = [
        "Portfolio Company",
        "Acquisition Financial Date",
        "Fund Name (GP)",
        "Fund Currency",
        "Cross-Fund Investment",
        "Country (HQ)",
        "Region of majority operations",
        "Kam Vertical",
        "Vertical Description",
        "Company Currency",
        "Investment strategy",
        "Instrument type",
        "Public/ Private",
        "Purchase Process",
        "Purchase Type",
        "Deal Role",
        "Seller Type",
        "Date Final Exit",
        "Exit type",
        "Fund Position Status",
        "Date Financials",
        "Date Investment",
        "Date IPO",
        "Total Investment Cost ($MM)",
        "Gross Proceeds Distributed to Fund",
        "Current Fair Value ($MM)",
        "Total Value ($MM)",
        "MOIC (Gross)",
        "IRR (Gross)",
        "Equity (Entry) ($MM)",
        "Kam Equity (Entry) ($MM)",
        "Net Debt (Entry) ($MM)",
        "Enterprise Value (Entry) ($MM)",
        "Revenue (Entry) ($MM)",
        "EBITDA (Entry) ($MM)",
        "Multiple (Entry)",
        "M&A Number of Transactions (Net)",
        "Most Recently Available Financial Statements",
        "Equity (Exit/Current) ($MM)",
        "Net Debt (Exit/Current) ($MM)",
        "Enterprise Value (Exit/Current) ($MM)",
        "Revenue (Exit/Current) ($MM)",
        "EBITDA (Exit/Current) ($MM)",
        "Multiple (Exit/Current) ($MM)",
        "Multiple Methodology (Entry)",
        "Multiple Methodology (Exit/Current) ($MM)",
        "Kam Equity (Entry)",
        "Kam Equity Ownership Fund (Exit/Current)",
        "Co-Invest Equity Ownership (Entry)",
        "Co-Invest Equity Ownership (Exit/Current)",
        "Lender Equity Ownership (Entry)",
        "Lender Equity Ownership (Exit/Current)",
        "Mgmt./Seller Equity Ownership (Entry)",
        "Mgmt./Seller Equity Ownership (Exit/Current)",
        "Post Rollover New Majority Owner(s) Ownership (Current)",
    ]

    portfolio_instructions = [
        "Required: Yes — Text. If no data: N/A not allowed; provide a unique company name.",
        "Required: Yes — Date (YYYY-MM-DD or Excel date). If no data: cannot be blank; used for filters, charts, and holding period.",
        "Required: No — Text. If no data: leave blank.",
        "Required: No — Text (e.g., USD, EUR). If no data: leave blank.",
        "Required: No — Text Yes/No. If no data: leave blank.",
        "Required: No — Text country. If no data: leave blank.",
        "Required: No — Text region. If no data: leave blank.",
        "Required: No — Text sector/vertical. If no data: leave blank.",
        "Required: No — Text. If no data: leave blank.",
        "Required: No — Text currency code. If no data: leave blank.",
        "Required: No — Text strategy. If no data: leave blank.",
        "Required: No — Text instrument type. If no data: leave blank.",
        "Required: No — Text Public/Private. If no data: leave blank.",
        "Required: No — Text purchase process. If no data: leave blank.",
        "Required: No — Text purchase type. If no data: leave blank.",
        "Required: No — Text deal role. If no data: leave blank.",
        "Required: No — Text seller type. If no data: leave blank.",
        "Required: No — Date (final exit if realized). If not realized: leave blank.",
        "Required: No — Text exit type. If no data: leave blank.",
        "Required: No — Text status (Realized/Unrealized/Partial). If no data: leave blank.",
        "Required: No — Date (most recent financials). If no data: leave blank.",
        "Required: No — Date (investment). If no data: leave blank.",
        "Required: No — Date (IPO if applicable). If no data: leave blank.",
        "Required: Yes — Number $MM invested. If no data: leave blank; deal excluded from track record.",
        "Required: Yes — Number $MM gross proceeds. If no data: leave blank; treat as 0 if truly none.",
        "Required: Yes — Number $MM current fair value (NAV). If no data: leave blank; treat as 0 if truly none.",
        "Required: No — Number $MM total value (Proceeds + NAV). If no data: leave blank (app can compute).",
        "Required: No — Number gross MOIC (e.g., 1.8). If no data: leave blank (app can compute).",
        "Required: No — Percent gross IRR (0.18 or 18%). If no data: leave blank.",
        "Required: No — Number $MM equity at entry. If no data: leave blank.",
        "Required: No — Number $MM Kam equity at entry. If no data: leave blank.",
        "Required: Yes — Number $MM net debt at entry. If no data: leave blank; deal excluded from value creation.",
        "Required: Yes — Number $MM enterprise value at entry. If no data: leave blank; deal excluded from value creation.",
        "Required: Yes — Number $MM revenue at entry. If no data: leave blank; deal excluded from value creation.",
        "Required: Yes — Number $MM EBITDA at entry. If no data: leave blank; deal excluded from value creation.",
        "Required: No — Number Entry TEV/EBITDA multiple. If no data: leave blank (app can compute).",
        "Required: No — Number net M&A transactions. If no data: leave blank.",
        "Required: No — Date most recent financial statements. If no data: leave blank.",
        "Required: No — Number $MM equity at exit/current. If no data: leave blank.",
        "Required: Yes — Number $MM net debt at exit/current. If no data: leave blank; deal excluded from value creation.",
        "Required: Yes — Number $MM enterprise value at exit/current. If no data: leave blank; deal excluded from value creation.",
        "Required: Yes — Number $MM revenue at exit/current. If no data: leave blank; deal excluded from value creation.",
        "Required: Yes — Number $MM EBITDA at exit/current. If no data: leave blank; deal excluded from value creation.",
        "Required: No — Number Exit/Current TEV/EBITDA multiple. If no data: leave blank (app can compute).",
        "Required: No — Text multiple methodology (entry). If no data: leave blank.",
        "Required: No — Text multiple methodology (exit/current). If no data: leave blank.",
        "Required: No — Number $MM Kam equity (entry). If no data: leave blank.",
        "Required: No — Percent Kam equity ownership fund (exit/current). If no data: leave blank.",
        "Required: No — Percent Co-invest equity ownership (entry). If no data: leave blank.",
        "Required: No — Percent Co-invest equity ownership (exit/current). If no data: leave blank.",
        "Required: No — Percent Lender equity ownership (entry). If no data: leave blank.",
        "Required: No — Percent Lender equity ownership (exit/current). If no data: leave blank.",
        "Required: No — Percent Mgmt./Seller equity ownership (entry). If no data: leave blank.",
        "Required: No — Percent Mgmt./Seller equity ownership (exit/current). If no data: leave blank.",
        "Required: No — Percent post-rollover new majority owner(s) ownership (current). If no data: leave blank.",
    ]

    # Funds Net Returns headers
    funds_headers = ["Fund", "Fund Size", "Vintage Year", "Net IRR", "Net TVPI", "Net DPI"]
    funds_instructions = [
        "Text: Fund name",
        "Number: Fund size ($MM)",
        "Integer: Vintage year (e.g., 2019)",
        "Percent: Net IRR (0.18 or 18%)",
        "Multiple: Net TVPI (e.g., 1.8)",
        "Multiple: Net DPI (e.g., 0.9)",
    ]

    # Benchmarks headers
    bench_headers = ["Vintage Year", "Metric", "Top5%", "Upper Quartile", "Median", "Lower Quartile"]
    bench_instructions = [
        "Integer: Vintage year (e.g., 2019)",
        "Text: One of 'Net IRR', 'Net TVPI', 'Net DPI'",
        "Threshold: Top 5% value (percent or multiple)",
        "Threshold: Upper quartile (percent or multiple)",
        "Threshold: Median (percent or multiple)",
        "Threshold: Lower quartile (percent or multiple)",
    ]

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # Portfolio Metrics sheet
        pd.DataFrame(columns=portfolio_headers).to_excel(
            writer, sheet_name="Portfolio Metrics", index=False, startrow=1
        )
        ws = writer.sheets["Portfolio Metrics"]
        for idx, instr in enumerate(portfolio_instructions, start=1):
            ws.cell(row=1, column=idx, value=instr)
        # Autosize a bit
        for i, hdr in enumerate(portfolio_headers, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = max(14, min(50, len(str(hdr)) + 4))

        # Funds Net Returns sheet
        pd.DataFrame(columns=funds_headers).to_excel(
            writer, sheet_name="Funds Net Returns", index=False, startrow=1
        )
        ws2 = writer.sheets["Funds Net Returns"]
        for idx, instr in enumerate(funds_instructions, start=1):
            ws2.cell(row=1, column=idx, value=instr)
        for i, hdr in enumerate(funds_headers, start=1):
            col_letter = get_column_letter(i)
            ws2.column_dimensions[col_letter].width = max(12, min(30, len(str(hdr)) + 4))

        # Benchmarks sheet
        pd.DataFrame(columns=bench_headers).to_excel(
            writer, sheet_name="Benchmarks", index=False, startrow=1
        )
        ws3 = writer.sheets["Benchmarks"]
        for idx, instr in enumerate(bench_instructions, start=1):
            ws3.cell(row=1, column=idx, value=instr)
        for i, hdr in enumerate(bench_headers, start=1):
            col_letter = get_column_letter(i)
            ws3.column_dimensions[col_letter].width = max(14, min(32, len(str(hdr)) + 4))

        # PME Cash Flows sheet (fourth sheet)
        pme_headers = ["Date", "Type", "Value", "Portfolio Company", "Fund"]
        pme_instructions = [
            "Required: Yes — Date (YYYY-MM-DD or Excel date).",
            "Required: Yes — Text: Capital Call / Distribution / NAV.",
            "Required: Yes — Number $MM; Calls negative, Distributions & NAV positive.",
            "Required: Yes — Text: Company name (matches Portfolio Metrics).",
            "Required: Yes — Text: Fund name (Column E).",
        ]
        pd.DataFrame(columns=pme_headers).to_excel(
            writer, sheet_name="PME Cash Flows", index=False, startrow=1
        )
        ws4 = writer.sheets["PME Cash Flows"]
        for idx, instr in enumerate(pme_instructions, start=1):
            ws4.cell(row=1, column=idx, value=instr)
        for i, hdr in enumerate(pme_headers, start=1):
            col_letter = get_column_letter(i)
            ws4.column_dimensions[col_letter].width = max(14, min(34, len(str(hdr)) + 4))

        # Fund Manager Profile (fifth sheet)
        mgr_headers = [
            "Firm Name",
            "Headquarters",
            "AUM ($MM)",
            "Strategy Focus",
            "Region Focus",
            "Year Founded",
            "Team Size",
            "Contact Email",
            "Website",
            "Description",
        ]
        mgr_instructions = [
            "Required: Yes — Firm name.",
            "Optional — City, Country.",
            "Optional — Total AUM in $MM.",
            "Optional — Primary strategy (e.g., Buyout, Growth).",
            "Optional — Primary regions.",
            "Optional — Year founded.",
            "Optional — Team size.",
            "Optional — Contact email.",
            "Optional — Firm website URL.",
            "Optional — Short description/bio.",
        ]
        pd.DataFrame(columns=mgr_headers).to_excel(
            writer, sheet_name="Fund Manager Profile", index=False, startrow=1
        )
        ws5 = writer.sheets["Fund Manager Profile"]
        for idx, instr in enumerate(mgr_instructions, start=1):
            ws5.cell(row=1, column=idx, value=instr)
        for i, hdr in enumerate(mgr_headers, start=1):
            col_letter = get_column_letter(i)
            ws5.column_dimensions[col_letter].width = max(16, min(50, len(str(hdr)) + 6))

    return bio.getvalue()


def _maybe_render_manager_banner(manager_profile: Optional[Dict[str, object]]) -> None:
    try:
        name = None
        if manager_profile:
            for key in ("firm_name", "name"):
                val = manager_profile.get(key)
                if val is not None and pd.notna(val):
                    name = str(val)
                    break
        if not name:
            name = st.session_state.get("firm_name")
        if name:
            st.markdown(
                f"""
                <style>
                .mgr-banner {{
                    background: #f5f7fb;
                    border: 1px solid #e4e8f0;
                    border-radius: 6px;
                    padding: 6px 10px;
                    margin-bottom: 8px;
                    font-weight: 600;
                }}
                </style>
                <div class=\"mgr-banner\">Fund Manager: {name}</div>
                """,
                unsafe_allow_html=True,
            )
    except Exception:
        pass

